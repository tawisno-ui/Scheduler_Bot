import os
import requests
from datetime import datetime
from openpyxl import load_workbook

PAGE_ACCESS_TOKEN = os.environ["PAGE_ACCESS_TOKEN"]
RECIPIENT_ID = os.environ["RECIPIENT_ID"]

SKIP_LABELS = {"BREAK TIME", "LUNCH BREAK"}

def send_message(text):
    url = "https://graph.facebook.com/v18.0/me/messages"
    params = {"access_token": PAGE_ACCESS_TOKEN}
    payload = {
        "recipient": {"id": RECIPIENT_ID},
        "message": {"text": text}
    }
    response = requests.post(url, params=params, json=payload)
    response.raise_for_status()
    print("Message sent successfully!")

def get_schedule():
    today = datetime.now().strftime("%A")
    print(f"Today is: {today}")

    wb = load_workbook("schedule.xlsx")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    time_slots = rows[1][1:]

    today_row = None
    for row in rows[2:]:
        if row[0] and str(row[0]).strip().title() == today:
            today_row = row[1:]
            break

    if today_row is None:
        return f"📌 No classes today ({today})! Enjoy your free day. 🎉"

    classes = [
        (time_slots[i], today_row[i])
        for i in range(len(time_slots))
        if i < len(today_row)
        and today_row[i]
        and str(today_row[i]).strip().upper() not in SKIP_LABELS
    ]

    if not classes:
        return f"📌 No classes today ({today})! Enjoy your free day. 🎉"

    lines = [f"📚 Schedule for {today}:\n"]
    for time_slot, subject in classes:
        lines.append(f"🕐 {time_slot}  —  {subject}")

    return "\n".join(lines)

if __name__ == "__main__":
    message = get_schedule()
    print(message)
    send_message(message)
